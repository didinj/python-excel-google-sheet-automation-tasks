import pandas as pd
import gspread
from openpyxl.styles import Font
from openpyxl import load_workbook
from gspread_dataframe import set_with_dataframe, get_as_dataframe
from oauth2client.service_account import ServiceAccountCredentials
import logging

# Setup logging
logging.basicConfig(
    filename="automation.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Google Sheets auth setup
def authorize_google_sheets():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive"
    ]
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
        client = gspread.authorize(creds)
        logging.info("Google Sheets authorization successful.")
        return client
    except Exception as e:
        logging.error(f"Google Sheets authorization failed: {e}")
        raise

# Read Excel
def read_excel(file_path):
    try:
        df = pd.read_excel(file_path)
        logging.info(f"Excel file '{file_path}' loaded.")
        return df
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        raise

# Write to Excel
def write_excel(df, file_path):
    try:
        df.to_excel(file_path, index=False)
        logging.info(f"Data written to Excel file '{file_path}'.")
    except Exception as e:
        logging.error(f"Error writing to Excel: {e}")
        raise

# Style Excel headers
def style_excel_headers(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(file_path)
        logging.info(f"Styled headers in '{file_path}'.")
    except Exception as e:
        logging.error(f"Error styling Excel: {e}")
        raise

# Push DataFrame to Google Sheets
def push_to_google_sheets(sheet_title, df):
    try:
        client = authorize_google_sheets()
        sheet = client.open(sheet_title).sheet1
        sheet.clear()
        set_with_dataframe(sheet, df)
        logging.info(f"Data pushed to Google Sheet: {sheet_title}.")
    except Exception as e:
        logging.error(f"Failed to update Google Sheet '{sheet_title}': {e}")
        raise

# Main script: Example pipeline
def main():
    try:
        df = pd.DataFrame({
            "Name": ["Alice", "Bob", "Charlie"],
            "Score": [85, 92, 78]
        })

        excel_file = "students.xlsx"
        sheet_title = "Students Report"

        write_excel(df, excel_file)
        style_excel_headers(excel_file)

        # Push to Google Sheets
        push_to_google_sheets(sheet_title, df)

    except Exception as e:
        logging.critical(f"Script failed: {e}")

if __name__ == "__main__":
    main()